import os
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from typing import List, Dict

from botocore.exceptions import ClientError
from ics import Calendar, Event
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from backend.core import Config
from backend.core.extensions import s3_client


def save_image(file: FileStorage, subfolder: str = "") -> str:
    """
    Сохраняет загруженное изображение в указанную папку проекта и возвращает относительный путь.

    :param file: объект загруженного файла (FileStorage)
    :param subfolder: подкаталог внутри папки загрузок
    :return: относительный путь к сохранённому файлу (например, 'media/uploads/news/filename.png')
    """
    original_filename = secure_filename(file.filename)
    name, ext = os.path.splitext(original_filename)
    unique_suffix = uuid.uuid4().hex
    filename = f"{name}_{unique_suffix}{ext}"

    # Ключ объекта в бакете
    key = f"{subfolder}/{filename}" if subfolder else filename

    # Загрузка файла в бакет
    s3_client.upload_fileobj(
        file,
        Config.BUCKET_NAME,
        key,
        ExtraArgs={"ACL": "public-read"}  # чтобы файл был публично доступен
    )

    # Возвращаем публичный URL
    return f"/media/uploads/{key}"


def remove_file_if_exists(file_path: str, use_bucket: bool = True) -> None:
    """
    Удаляет файл локально или в Yandex Object Storage, если он существует.

    :param file_path: путь к файлу (для бакета — ключ объекта)
    :param use_bucket: True — удалять из бакета, False — локально
    """
    file_path = file_path.lstrip('/media/uploads/')
    if use_bucket:
        try:
            s3_client.delete_object(Bucket=Config.BUCKET_NAME, Key=file_path)
            print(f"Файл {file_path} удалён из бакета {Config.BUCKET_NAME}")
        except ClientError as e:
            print(f"Ошибка при удалении файла из бакета {file_path}: {e}")
    else:
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
                print(f"Файл {file_path} удалён локально")
            except Exception as e:
                print(f"Ошибка при удалении локального файла {file_path}: {e}")


def format_datetime(value: datetime) -> str:
    """
    Форматирует datetime в строку "дд.мм.гггг чч:мм".

    :param value: объект datetime
    :return: строковое представление даты и времени
    """
    if isinstance(value, datetime):
        return value.strftime('%d.%m.%Y %H:%M')
    return str(value)


def generate_reservations_csv(reservations: List[Dict]) -> bytes:
    """
    Генерирует Excel-файл с отменёнными бронированиями.

    :param reservations: список словарей с данными бронирований
    :return: байты Excel-файла
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Отменённые бронирования"

    headers = [
        ('reservation_id', 'ID бронирования'),
        ('full_name', 'ФИО'),
        ('email', 'Электронная почта'),
        ('phone_number', 'Телефон'),
        ('participants_count', 'Количество участников'),
        ('booked_at', 'Дата бронирования'),
        ('session_datetime', 'Время сессии'),
        ('excursion_title', 'Название экскурсии'),
        ('place', 'Место экскурсии'),
        ('total_cost', 'Общая стоимость'),
        ('is_paid', 'Оплачена'),
        ('is_cancelled', 'Отменена'),
    ]
    ws.append([h[1] for h in headers])

    for r in reservations:
        ws.append([
            str(r.get('reservation_id', '')),
            str(r.get('full_name', '')),
            str(r.get('email', '')),
            str(r.get('phone_number', '')),
            str(r.get('participants_count', '')),
            format_datetime(r.get('booked_at', '')),
            format_datetime(r.get('session_datetime', '')),
            str(r.get('excursion_title', '')),
            str(r.get('place', '')),
            str(r.get('total_cost', '')),
            'Да' if r.get('is_paid') else 'Нет',
            'Да' if r.get('is_cancelled') else 'Нет',
        ])

    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def create_ical_from_reservation(reservation) -> bytes:
    """
    Создает iCal-файл для бронирования экскурсии.

    :param reservation: объект Reservation с привязанной сессией и мероприятием
    :return: байты iCal-файла
    """
    c = Calendar()
    e = Event()

    e.name = f"Событие: {reservation.session.event.title}"
    e.begin = reservation.session.start_datetime
    duration_minutes = getattr(reservation.session.event, 'duration', 60)
    e.duration = timedelta(minutes=duration_minutes)

    e.location = reservation.session.event.place or ""
    e.description = f"Бронирование экскурсии. Участников: {reservation.participants_count}"

    c.events.add(e)
    return c.serialize().encode('utf-8')
